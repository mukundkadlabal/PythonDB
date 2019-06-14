# -*- coding: utf-8 -*-
"""
Created on Fri Jun 17 20:49:36 2019

@author: 91961
"""

import tkinter
import datetime
import re
from functools import partial
from dbcreate import *
from tk2 import *
from openpyxl import Workbook
from sqlalchemy import Column,Integer, String, Float,Date, create_engine, exc, orm, MetaData, Table
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
meta = MetaData()
def addCourse():
    addCourseWin = tkinter.Tk()
    addCourseWin.title("Add a course")

 
    global addCourseLabel
 
    addCourseLabel = tkinter.StringVar()
 
    global addCourseEntry
   
    tkinter.Label(addCourseWin, text="Course : ").pack()
    
    addCourseEntry = tkinter.Entry(addCourseWin)
    addCourseEntry.pack()
    
    #print(addCourseEntry.get()
    addCourseLabel = addCourseEntry
    
    tkinter.Label(addCourseWin, text="").pack()
    
    tkinter.Button(addCourseWin, text="Submit", width=10, height=1, command = checkAdd).pack()
    
def checkAdd():
    print(addCourseLabel.get())
    addcoursefac(addCourseLabel.get(),usernames)

def viewCourse():
    viewWin = tkinter.Tk()
    viewWin.title("Your courses")
    courseList = getcoursesfac(usernames)
    for i in courseList:
        var = tkinter.Button(viewWin, text=i, command=partial(facAttendance,i), bg='gray', fg='white', height="2", width="25")
        var.pack()

def submitBtn(i):

    print(fromDate.get(),toDate.get())
    date1 = fromDate.get().split('/')
    date2 = toDate.get().split('/')
    
    x = datetime.date(int(date1[2]),int(date1[1]),int(date1[0]))
    y = datetime.date(int(date2[2]),int(date2[1]),int(date2[0]))
    
    madattendance(x,i,y) #x - the from date and y - to date . course = name of course whose attendance we're seeing

def facAttendance(course):
    attWin = tkinter.Tk()
    attWin.title(course)
    
    global fromDate
    global toDate
    
    global tbFrom
    global tbTo
    
    fromDate = tkinter.StringVar()
    toDate = tkinter.StringVar()

    fDt = tkinter.Label(attWin, text = "Enter front date in the following textbox as (d/m/y)")
    fDt.pack()
    
    tbFrom = tkinter.Entry(attWin)
    tbFrom.pack()
    fromDate = tbFrom
    
    tDt = tkinter.Label(attWin, text = "Enter to date in the following textbox as (d/m/y)")
    tDt.pack()
    
    tbTo = tkinter.Entry(attWin)
    tbTo.pack()
    toDate = tbTo
    
    submit = tkinter.Button(attWin, text='Submit',
        command=partial(submitBtn,course), bg='red', fg='white', height="2", width="25")
    submit.pack()
    
def excel():    
	engine=create_engine("sqlite:///as1.db")
	engine.connect()
	session= sessionmaker(bind=engine)()    
	meta = MetaData()    
	meta.create_all(engine)
	Base=declarative_base()

	wb = Workbook()
	ws = wb.active

	sheets=[]
	var = 1
	 
	result=[r for r in session.query(Perc).all()]
	for i in result:
		sheets = ws.cell(row=var, column=1)
		sheets.value = i.studname	
		sheets = ws.cell(row=var, column=2)
		sheets.value = i.course	

		sheets = ws.cell(row=var, column=3)
		sheets.value = i.perc	

		var+=1

	wb.save('students.xlsx')

def email():
	engine=create_engine("sqlite:///as1.db")
	engine.connect()
	session= sessionmaker(bind=engine)()    
	meta = MetaData()    
	meta.create_all(engine)
	Base=declarative_base()
	wb = Workbook()
	ws = wb.active
	sheets=[]
	var = 1
	result=[r for r in session.query(Perc).all()]
	for i in result:
		sheets = ws.cell(row=var, column=1)
		sheets.value = i.studname	
		sheets = ws.cell(row=var, column=2)
		sheets.value = i.course	

		sheets = ws.cell(row=var, column=3)
		sheets.value = i.perc	

		var+=1

	wb.save('studentseml.xlsx')
	# Import smtplib for the actual sending function
	import smtplib
	# For guessing MIME type
	import mimetypes
	# Import the email modules we'll need
	import email
	import email.mime.application
	import email.mime.multipart as mm
	import email.mime.text as tt
	# Create a text/plain message
	msg = mm.MIMEMultipart()
	msg['Subject'] = 'Greetings'
	msg['From'] = 'abcd@gmail.com'
	msg['To'] = 'abcd@gmail.com'
	# The main body is just another attachment
	body = tt.MIMEText("""Hello, how are you? I am fine.
	This is a rather nice letter, don't you think?""")
	msg.attach(body)
	# PDF attachment
	filename=r'/home/rithvik/Serious_python/Applications/newExcel.xlsx'
	fp=open(filename,'rb')
	att = email.mime.application.MIMEApplication(fp.read(),_subtype="xlsx")
	fp.close()
	att.add_header('Content-Disposition','attachment',filename=filename)
	msg.attach(att)
	# send via Gmail server
	# NOTE: my ISP, Centurylink, seems to be automatically rewriting
	# port 25 packets to be port 587 and it is trashing port 587 packets.
	# So, I use the default port 25, but I authenticate.
	s = smtplib.SMTP("smtp.gmail.com", 587)
	s.starttls()
	s.login('abcd@gmail.com','abcd')
	s.sendmail('abcd@gmail.com',['abcd@gmail.com'], msg.as_string())
	s.quit()


def sendToFTP():
	engine=create_engine("sqlite:///as1.db")
	engine.connect()
	session= sessionmaker(bind=engine)()    
	meta = MetaData()    
	meta.create_all(engine)
	Base=declarative_base()
	wb = Workbook()
	ws = wb.active
	sheets=[]
	var = 1
	result=[r for r in session.query(Perc).all()]
	for i in result:
		sheets = ws.cell(row=var, column=1)
		sheets.value = i.studname	
		sheets = ws.cell(row=var, column=2)
		sheets.value = i.course	

		sheets = ws.cell(row=var, column=3)
		sheets.value = i.perc	

		var+=1

	wb.save('studentseml.xlsx')

	import pysftp as sftp

	s=sftp.Connection(host='127.0.0.1',username='rithvik',password='ritzzz')
	remotepath="home/Serious_python/Applications/studentseml.xlsx"
	localpath="home/Serious_python/Applications/ftp/studentseml.xlsx"
	s.put(localpath,removepath)
	s.close()
      
def faculty(usernames1):
    global usernames
    usernames=usernames1
    mainWin = tkinter.Tk()
    mainWin.title("Faculty Window")
    mainWin.geometry("300x300")
    
    hello = tkinter.Label(mainWin, text='Faculty')
    hello.pack()
    
    viewBtn = tkinter.Button(mainWin, text='View courses',
        command=viewCourse, bg='red', fg='white', height="2", width="25")
    viewBtn.pack()
    
    courses = tkinter.Button(mainWin, text='Add course',
        command=addCourse, bg='red', fg='white', height="2", width="25")
    courses.pack()
    
    courses = tkinter.Button(mainWin, text='Excel to FTP',
        command=sendToFTP, bg='red', fg='white', height="2", width="25")
    courses.pack()
    
    courses = tkinter.Button(mainWin, text='Email excel',
        command=email, bg='red', fg='white', height="2", width="25")
    courses.pack()
    

    
#=================================student part=================================
    
    
def stuAttendance():
    stuAttWin = tkinter.Tk()
    stuAttWin.title("Viewing attendance")
    
    studentAttendance = getpercrows(usernames)
    #studentAttendance = [('cse',85),('python',75)]
    cnt=1
    for i in studentAttendance:
        tkinter.Label(stuAttWin, text=i[1]).grid(row=cnt,column=1)
        tkinter.Label(stuAttWin, text=i[2]).grid(row=cnt,column=2)
        cnt+=1


def stuCourse():
    stuCourseWin = tkinter.Tk()
    stuCourseWin.title("courses")
    allCourses = getallcourses() #get all courses here and remove the next list   
    for i in allCourses:
        cBtn = tkinter.Button(stuCourseWin, text=i,command=partial(addStuCourse, i), bg='lightblue', fg='black', height="2", width="25")
        cBtn.pack()

def addStuCourse():
    addstuCourseWin = tkinter.Tk()
    addstuCourseWin.title("Add a course")

    global addstuCourseLabel
 
    addstuCourseLabel = tkinter.StringVar()
 
    global addstuCourseEntry
   
    tkinter.Label(addstuCourseWin, text="Course : ").pack()
    
    addstuCourseEntry = tkinter.Entry(addstuCourseWin)
    addstuCourseEntry.pack()
    #print(addCourseEntry.get()
    addstuCourseLabel = addstuCourseEntry
    
    tkinter.Label(addstuCourseWin, text="").pack()
    
    tkinter.Button(addstuCourseWin, text="Submit", width=10, height=1, command = checkstuAdd).pack()


def checkstuAdd():
    print(addstuCourseLabel.get())
    if (addstuCourseLabel.get() in getallcourses()):
    	addcoursestud(usernames,addstuCourseLabel.get()) #update daatabase function here
    else:
    	print("No course yet")


def searchFunc():
    pass    

def student(usernames1):
    global usernames
    usernames=usernames1
    stuMain = tkinter.Tk()
    stuMain.title("Student Window")
    
    stu = tkinter.Label(stuMain, text='Student')
    stu.pack()
    
    #cBtn = tkinter.Button(stuMain, text='Select Courses',command=stuCourse, bg='red', fg='white', height="2", width="25")
    #cBtn.pack()
    
    csBtn = tkinter.Button(stuMain, text='Add Courses',command=addStuCourse, bg='red', fg='white', height="2", width="25")
    csBtn.pack()

    att = tkinter.Button(stuMain, text='My Attendance',command=stuAttendance, bg='red', fg='white', height="2", width="25")
    att.pack()
    
    search = tkinter.Button(stuMain, text='Search', command=searchFunc, bg='red', fg='white', height='2', width='25')
    search.pack()


#to rithvik : add that thing from database isFaculty == 0 => then call student() or else faculty() above